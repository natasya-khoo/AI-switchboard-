"""
Excel Export Utilities
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import io
import pandas as pd

class ExcelExporter:
    """Handle Excel export operations"""
    
    def __init__(self):
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.title_font = Font(bold=True, size=14)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def export_for_erp(self, project, bom_items, company_name):
        """
        Export BOM in format suitable for manual ERP entry (sosopoih format)
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "BOM for ERP Entry"
        
        # Header section
        ws['A1'] = company_name.upper()
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:K1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws['A2'] = "BILL OF MATERIALS - ERP ENTRY FORMAT"
        ws['A2'].font = self.title_font
        ws.merge_cells('A2:K2')
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Project details
        row = 4
        ws[f'A{row}'] = "Project Code:"
        ws[f'B{row}'] = project['project_code']
        ws[f'B{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "Project Name:"
        ws[f'B{row}'] = project['project_name']
        
        row += 1
        ws[f'A{row}'] = "Client:"
        ws[f'B{row}'] = project.get('client_name', '')
        
        row += 1
        ws[f'A{row}'] = "Date:"
        ws[f'B{row}'] = datetime.now().strftime('%Y-%m-%d')
        
        # Column headers (matching sosopoih structure)
        row = 9
        headers = [
            'SEQ', 'ITEMNAME', 'ITEMDESC', 'ITDESC2', 'ITDESC3', 'ITDESC4',
            'ITCLASS', 'QTY', 'UNITPRC', 'MARKUP%', 'NOTES'
        ]
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        # Data rows
        for idx, item in enumerate(bom_items, start=1):
            row += 1
            
            data = [
                idx,                                    # SEQ
                item['itemname'],                       # ITEMNAME
                item.get('itemdesc', ''),              # ITEMDESC
                item.get('itdesc2', ''),               # ITDESC2
                item.get('itdesc3', ''),               # ITDESC3
                item.get('itdesc4', ''),               # ITDESC4
                item['itclass'],                        # ITCLASS
                item['qty'],                            # QTY
                float(item['unit_price']),             # UNITPRC
                float(item['markup_pct']),             # MARKUP%
                item.get('notes', '')                   # NOTES
            ]
            
            for col_idx, value in enumerate(data, start=1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.border = self.border
                
                # Format numbers
                if col_idx == 9:  # Unit price
                    cell.number_format = '$#,##0.00'
                elif col_idx == 10:  # Markup
                    cell.number_format = '0.00"%"'
        
        # Summary section
        row += 2
        
        summary_data = [
            ('Materials Total:', project['total_materials_cost']),
            ('Labor Cost:', project['total_labor_cost']),
            ('Subtotal:', project['total_materials_cost'] + project['total_labor_cost']),
            ('Markup:', project['total_markup']),
            ('GRAND TOTAL:', project['grand_total'])
        ]
        
        for label, value in summary_data:
            ws[f'I{row}'] = label
            ws[f'I{row}'].font = Font(bold=True) if 'TOTAL' in label else Font()
            ws[f'I{row}'].alignment = Alignment(horizontal='right')
            
            ws[f'J{row}'] = value
            ws[f'J{row}'].number_format = '$#,##0.00'
            ws[f'J{row}'].font = Font(bold=True) if 'TOTAL' in label else Font()
            
            row += 1
        
        # Instructions
        row += 2
        ws[f'A{row}'] = "INSTRUCTIONS FOR ERP ENTRY:"
        ws[f'A{row}'].font = Font(bold=True, color="FF0000")
        
        row += 1
        instructions = [
            "1. Copy data from columns B to K (ITEMNAME to NOTES)",
            "2. Paste into sosopoih table in ERP system",
            "3. Ensure PARENTKY field is set to correct Sales Order number",
            "4. Verify all prices and quantities before finalizing",
            "5. Update SINDEX field as per your ERP trigger logic"
        ]
        
        for instruction in instructions:
            ws[f'A{row}'] = instruction
            ws[f'A{row}'].font = Font(italic=True)
            row += 1
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            
            for cell in ws[column_letter]:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    def export_detailed(self, project, bom_items, company_name, company_address):
        """
        Export detailed BOM with full specifications and cost breakdown
        """
        wb = openpyxl.Workbook()
        
        # Main BOM sheet
        ws1 = wb.active
        ws1.title = "Bill of Materials"
        
        # Header
        ws1['A1'] = company_name.upper()
        ws1['A1'].font = Font(size=18, bold=True)
        ws1.merge_cells('A1:H1')
        
        if company_address:
            ws1['A2'] = company_address
            ws1.merge_cells('A2:H2')
            header_row = 4
        else:
            header_row = 3
        
        # Project info
        ws1[f'A{header_row}'] = "DETAILED BILL OF MATERIALS"
        ws1[f'A{header_row}'].font = Font(size=14, bold=True)
        ws1.merge_cells(f'A{header_row}:H{header_row}')
        
        info_row = header_row + 2
        
        project_info = [
            ('Project Code:', project['project_code']),
            ('Project Name:', project['project_name']),
            ('Client:', project.get('client_name', 'N/A')),
            ('Date:', datetime.now().strftime('%Y-%m-%d')),
            ('Status:', project['status'].title())
        ]
        
        for label, value in project_info:
            ws1[f'A{info_row}'] = label
            ws1[f'A{info_row}'].font = Font(bold=True)
            ws1[f'B{info_row}'] = value
            info_row += 1
        
        # BOM table
        table_row = info_row + 2
        
        headers = [
            'Line', 'Item Name', 'Class', 'Manufacturer', 'Model',
            'Qty', 'Unit Price', 'Line Total'
        ]
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws1.cell(row=table_row, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        # Data
        for item in bom_items:
            table_row += 1
            
            data = [
                item.get('line_sequence', ''),
                item['itemname'],
                item['itclass'],
                item.get('manufacturer', ''),
                item.get('model_number', ''),
                item['qty'],
                float(item['unit_price']),
                float(item['line_total'])
            ]
            
            for col_idx, value in enumerate(data, start=1):
                cell = ws1.cell(row=table_row, column=col_idx, value=value)
                cell.border = self.border
                
                if col_idx in [7, 8]:  # Price columns
                    cell.number_format = '$#,##0.00'
        
        # Cost summary
        table_row += 2
        
        ws1[f'F{table_row}'] = "Materials:"
        ws1[f'F{table_row}'].font = Font(bold=True)
        ws1[f'H{table_row}'] = project['total_materials_cost']
        ws1[f'H{table_row}'].number_format = '$#,##0.00'
        
        table_row += 1
        ws1[f'F{table_row}'] = f"Labor ({project['total_labor_hours']:.1f} hours):"
        ws1[f'F{table_row}'].font = Font(bold=True)
        ws1[f'H{table_row}'] = project['total_labor_cost']
        ws1[f'H{table_row}'].number_format = '$#,##0.00'
        
        table_row += 1
        ws1[f'F{table_row}'] = f"Markup ({project['default_markup_pct']:.1f}%):"
        ws1[f'F{table_row}'].font = Font(bold=True)
        ws1[f'H{table_row}'] = project['total_markup']
        ws1[f'H{table_row}'].number_format = '$#,##0.00'
        
        table_row += 1
        ws1[f'F{table_row}'] = "GRAND TOTAL:"
        ws1[f'F{table_row}'].font = Font(bold=True, size=12)
        ws1[f'H{table_row}'] = project['grand_total']
        ws1[f'H{table_row}'].number_format = '$#,##0.00'
        ws1[f'H{table_row}'].font = Font(bold=True, size=12)
        
        # Component Details sheet
        ws2 = wb.create_sheet("Component Details")
        
        ws2['A1'] = "COMPONENT SPECIFICATIONS"
        ws2['A1'].font = Font(size=14, bold=True)
        ws2.merge_cells('A1:F1')
        
        detail_headers = [
            'Item Name', 'Description 1', 'Description 2',
            'Description 3', 'Description 4', 'Notes'
        ]
        
        for col_idx, header in enumerate(detail_headers, start=1):
            cell = ws2.cell(row=3, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        
        for idx, item in enumerate(bom_items, start=4):
            data = [
                item['itemname'],
                item.get('itemdesc', ''),
                item.get('itdesc2', ''),
                item.get('itdesc3', ''),
                item.get('itdesc4', ''),
                item.get('notes', '')
            ]
            
            for col_idx, value in enumerate(data, start=1):
                cell = ws2.cell(row=idx, column=col_idx, value=value)
                cell.border = self.border
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Auto-adjust columns
        for ws in [ws1, ws2]:
            for col in range(1, 10):
                column_letter = get_column_letter(col)
                max_length = 0
                
                for cell in ws[column_letter]:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 60)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    def export_csv(self, bom_items):
        """
        Export BOM as CSV
        """
        df = pd.DataFrame(bom_items)
        
        # Select columns
        columns = [
            'itemname', 'itclass', 'manufacturer', 'model_number',
            'qty', 'unit_price', 'markup_pct', 'line_total',
            'itemdesc', 'notes'
        ]
        
        df = df[[col for col in columns if col in df.columns]]
        
        # Export to CSV
        output = io.StringIO()
        df.to_csv(output, index=False)
        
        return output.getvalue().encode('utf-8')